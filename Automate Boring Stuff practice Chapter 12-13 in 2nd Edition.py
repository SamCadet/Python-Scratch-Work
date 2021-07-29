import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
from openpyxl.styles import Font

# openpyxl.cell is outdated, p. 270

# wb = openpyxl.load_workbook('example.xlsx')
# print(wb.sheetnames)
# sheet = wb['Sheet3']  # not wb.get_sheet_name() anymore p. 268
# print(sheet)
# print(type(sheet))
# print(sheet.title)
# another_sheet = wb.active  # not web.get_active_sheet() anymore p. 268
# print(sheet)
# print(another_sheet)
#
# sheet = wb['Sheet1']
# print(sheet['A1'])
# print(sheet['A1'].value)
# c = sheet['B1']
# print(c.value)
# print('Row ' + str(c.row) + ', Column ' +
#       str(c.column) + ' is ' + str(c.value))
#
# print('Cell ' + c.coordinate + ' is ' + c.value)
#
# print(sheet['C1'].value)
#
# print(sheet.cell(row=1, column=2))
# print(sheet.cell(row=1, column=2).value)
# for i in range(1, 8, 2):
#     print(i, sheet.cell(row=i, column=2).value)
#
# sheet = wb['Sheet1']
# print(sheet.max_row) # not sheet.get_highest_row anymore p. 269
# print(sheet.max_column) # not sheet.get_highest_column anymore p. 269

# print(get_column_letter(1))
# print(get_column_letter(2))
# print(get_column_letter(27))
# print(get_column_letter(900))

# print(get_column_letter(sheet.max_column))
# print(column_index_from_string('A'))
# print(column_index_from_string('AA'))

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb['Sheet1']
#
# (tuple(sheet['A1':'C3']))
#
# for row_of_cell_objects in sheet['A1':'C3']:
#     for cell_obj in row_of_cell_objects:
#         print(cell_obj.coordinate, cell_obj.value)
#     print('--- END OF ROW ---')

# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.active
# print(list(sheet.columns)[1])  # Get second column's cells.
#
# for cell_obj in list(sheet.columns)[1]:
#     print(cell_obj.value)

# wb = openpyxl.Workbook()  # create a blank workbook.
# print(wb.sheetnames)  # It starts with one sheet.
# sheet = wb.active
# print(sheet.title)
# sheet.title = 'Spam Bacon Eggs Sheet'  # Change title.
# print(wb.sheetnames)
# print(os.getcwd())
# wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.active
# sheet.title = 'Spam Spam Spam'
# wb.save('example_copy.xlsx')  # Save the workbook.

# wb = openpyxl.Workbook()
# print(wb.sheetnames)
# print(wb.create_sheet())  # Add a new sheet.
# print(wb.sheetnames)

# Create a new sheet at index 0.

# print(wb.create_sheet(index=0, title='First Sheet'))
#
# print(wb.sheetnames)
#
# print(wb.create_sheet(index=2, title='Middle Sheet'))
# print(wb.sheetnames)
#
# del wb['Middle Sheet']
# del wb['Sheet1']
# print(wb.sheetnames)

# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# sheet['A1'] = 'Ayo, myguy!'  # Edit the cell's value.
# print(sheet['A1'].value)


# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# Create a font.
# italic24Font = Font(name="Calibri", size=24, italic=True)
# sheet['A1'].font = italic24Font  # Apply the font to A1.
# sheet['A1'] = 'What\'s up wit\' it, whoadie!'
#
# font_obj1 = Font(name='Times New Roman', bold=True)
# sheet['A5'].font = font_obj1
# sheet['A5'] = 'Bold Times New Roman'
# font_obj2 = Font(size=24, italic=True)
# sheet['B9'].font = font_obj2
# sheet['B9'] = '24 pt Italic'
#
#
# wb.save('styles.xlsx')
# print('Changes saved, #MyGuy!')

# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 200
# sheet['A2'] = 300
# sheet['A3'] = '=SUM(A1:A2)' # Set the formula.
# wb.save('Write Formula.xlsx')

# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 'Tall row'
# sheet['B2'] = 'Wide column'
# Set the height and width:
# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions['B'].width = 20
# wb.save('dimensions.xlsx')
#
# print('Changes saved, #MyGuy')

# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.merge_cells('A1:D3')  # Merge all these cells.
# sheet['A1'] = 'Twelve cells merged together.'
# sheet.merge_cells('C5:D5')  # Merge these two cells.
# sheet['C5'] = 'Two merged cells.'
#
# wb.save('merged.xlsx')
# print('Changes saved, #MyGuy')

# wb = openpyxl.load_workbook('merged.xlsx')
# sheet = wb.active
# sheet.unmerge_cells('A1:D3')  # Split these cells up.
# sheet.unmerge_cells('C5:D5')
# wb.save('merged.xlsx')

# wb = openpyxl.load_workbook('Produce Sales.xlsx')
# sheet = wb.active
# sheet.freeze_panes = 'A2'  # Freeze the rows above A2.
# wb.save('Freeze Example.xlsx')
#
# print('Changes saved, #MyGuy')

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):  # create some data in column A
    sheet['A' + str(i)] = i

ref_obj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1,
                                   max_row=10)
series_obj = openpyxl.chart.Series(ref_obj, title='First series')
chart_obj = openpyxl.chart.BarChart()
chart_obj.title = 'My Chart'
chart_obj.append(series_obj)
sheet.add_chart(chart_obj, 'C5')
wb.save('Sample Chart.xlsx')

print('Changes saved, #MyGuy')
