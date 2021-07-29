#! python3

# sendDuesReminders.py - gotta remind the broke bois to pay their dues, dammit

# p. 432 - 434

# Step 1: Open the Excel File

import openpyxl
import smtplib
import sys

# Open the spreadsheet and get the latest dues status.

thatSheet = 'C:\\Users\\Sam\\PythonScripts\\duesRecords.xlsx'

wb = openpyxl.load_workbook(thatSheet)
sheet = wb['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol). value

# Step 2: Find All Unpaid Members

# Check each member's payment status.

unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Step 3: Send Customized Email Reminders

# Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
# Figure out why this won't run in command line, use app password in the meantime
smtpObj.login('clonlineorders23@gmail.com', sys.argv[1])

'''
pass it your email address and sys.argv[1], which will store
your password string. You’ll enter the password as a command line argument each time you run the
program, to avoid saving your password in your source code.
'''
# Send out reminder emails.

for name, email in unpaidMembers.items():
    body = '''Subject: %s dues unpaid.\nDear %s, \nRecords show that you have not
    paid dues for %s. Please make this payment as soon as possible. Thank you!''' % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail(
        'clonlineorders23@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' %
              (email, sendmailStatus))

smtpObj.quit()
